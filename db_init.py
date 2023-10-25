import pymongo
import openpyxl as xl
xlsx=xl.load_workbook("data/attendance.xlsx",data_only=True)
sheet=xlsx["I_CSE-B"]
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["Project-X"]
collection = db["studentDet"]
column=0
last_row = None
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=4, column=row).value is not None:
        last_row = row
reg='D'
roll='C'
name='E'
userlist=[]
ap='BA'
tp='BB'
AP = xl.utils.column_index_from_string(ap)
TP = xl.utils.column_index_from_string(tp)
for row in range(1, sheet.max_row + 1):
    reg_no = sheet[reg + str(row)].value
    roll_no = sheet[roll + str(row)].value
    _id=row
    namee=sheet[name + str(row)].value
    attended=sheet.cell(row,AP).value
    print(attended)
    total=sheet.cell(row,TP).value
    if reg_no and roll_no and _id and namee !=None and isinstance(reg_no, int) and isinstance(attended, int):
        perc=(attended/total)*100
        user_dict = {
            "_id": _id,
            "name": namee,
            "Roll No": roll_no,
            "Reg No": reg_no,
            "Yr": "II",
            "Age": 18,
            "Sem": 3,
            "DOB": "1/1/2001",
            "Email": "Sample@email.com",
            "Cls": "CSE-B",
            "phone": 9876543210,
            "attendance": perc,
            "atte_per":attended,
            "total_per":total
        }
        exist=collection.find_one({"Roll No": roll_no} )
        if exist:
            percc=exist["_id"]
            print(perc)
            print("skipped")
            update_query = {
            "$set": {
            "attendance": perc,
            "atte_per": attended,
            "total_per": total
                        }
                    }
            res=collection.update_one({"Roll No": roll_no},update_query)
            #collection.update_one({"Roll No": roll_no},{"$set":{"atte_per":attended}})
            #collection.update_one({"Roll No": roll_no},{"$set":{"total_per":total}})

            print(res)
        else:
            print("added")
            collection.insert_one(user_dict)
        #userlist.append(user_dict)  
print(userlist)
