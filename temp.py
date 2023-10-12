import pymongo
import openpyxl as xl
xlsx=xl.load_workbook("data/attendance.xlsx")
sheet=xlsx["I_CSE-B"]
print(sheet.cell(10,10).value)
column=0
last_row = None
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=4, column=row).value is not None:
        last_row = row
reg='D'
roll='C'
name='E'
userlist=[]
print(last_row)
for row in range(1, sheet.max_row + 1):
    reg_no = sheet[reg + str(row)].value
    roll_no = sheet[roll + str(row)].value
    _id=row
    namee=sheet[name + str(row)].value
    if reg_no and roll_no and _id and namee !=None and isinstance(reg_no, int):
        user_dict = {
            "_id": _id,
            "name": namee,
            "Roll No": roll_no,
            "Reg No": reg_no
        }
        userlist.append(user_dict)
print(userlist)
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["mydatabase"]
collection = db["studentusercred"]
collection.insert_many(userlist)