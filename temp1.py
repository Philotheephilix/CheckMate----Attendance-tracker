import pymongo
import openpyxl as xl
xlsx=xl.load_workbook("data/attendance.xlsx")
sheet=xlsx["I_CSE-B"]
column=0
last_row = None
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=4, column=row).value is not None:
        last_row = row
reg='D'
ap='BA'
tp='BB'
userlist=[]
print(last_row)
for row in range(1, sheet.max_row + 1):
    reg_no = sheet[reg + str(row)].value
    attended=sheet[ap + str(row)].value
    print(attended)
    total=sheet[tp + str(row)].value
    print(total)
    _id=row
    perc=(attended/total)*100
    if reg_no and _id !=None and isinstance(reg_no, int):
        user_dict = {
            "_id": _id,
            "Reg No": reg_no,
            "attendance": perc
        }
        userlist.append(user_dict)
print(userlist)
#client = pymongo.MongoClient("mongodb://localhost:27017/")
#db = client["mydatabase"]
#collection = db["studentusercred"]
#collection.insert_many(userlist)