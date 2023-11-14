import pandas as pd
import os
import csv
from flask import Flask, render_template, request, redirect, url_for, send_file
import pymongo
import xlsxwriter
from PIL import Image
import gridfs
import io
import base64
import openpyxl as xl
client = pymongo.MongoClient("mongodb+srv://maintainer_philix:mongotest@carbonpi.hiozz58.mongodb.net/?retryWrites=true&w=majority")

try:
    os.mkdir("report_data")
except:
    pass
for i in ["CSE","MECH","IT","ECE","EEE"]:
    try:
        os.mkdir("data/"+i)

    except:
        pass
for i in ["CSE","MECH","IT","ECE","EEE"]:
    for j in ["I","II","III","IV"]:
        try:
            os.mkdir("data/"+i+"/"+j)

        except:
            pass
app = Flask(__name__)
UPLOAD_FOLDER = 'static/profile_db'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER



def init_db(filename,yr):
    xlsx=xl.load_workbook(filename,data_only=True)
    sheet=xlsx.worksheets[0]
    db = client["Project-X"]
    collection = db["studentDet"]
    column=0
    last_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=4, column=row).value is not None:
            last_row = row
    reg='D'
    roll='C'
    name='E'
    userlist=[]
    ap='BA'
    tp='BB'
    AP = xl.utils.column_index_from_string(ap)
    TP = xl.utils.column_index_from_string(tp)
    for row in range(1, sheet.max_row + 1):
        reg_no = sheet[reg + str(row)].value
        roll_no = sheet[roll + str(row)].value
        _id=row
        namee=sheet[name + str(row)].value
        attended=sheet.cell(row,AP).value
        print(attended)
        total=sheet.cell(row,TP).value
        if reg_no and roll_no and _id and namee !=None and isinstance(reg_no, int) and isinstance(attended, int):
            perc=(attended/total)*100
            user_dict = {
                "_id": _id,
                "name": namee,
                "Roll No": roll_no,
                "Reg No": reg_no,
                "Yr": yr ,
                "Age": 18,
                "Sem": 3,
                "DOB": "1/1/2001",
                "Email": "Sample@email.com",
                "Cls": "CSE-B",
                "phone": 9876543210,
                "attendance": perc,
                "atte_per":attended,
                "total_per":total
            }
            exist=collection.find_one({"Roll No": roll_no} )
            if exist:
                percc=exist["_id"]
                print(perc)
                print("skipped")
                update_query = {
                "$set": {
                "attendance": perc,
                "atte_per": attended,
                "total_per": total
                            }
                        }
                res=collection.update_one({"Roll No": roll_no},update_query)
            else:
                collection.insert_one(user_dict) 
            print("all data added") 
        print(userlist)
    return "SUCCESS"     
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
try:
    os.mkdir("report_data")
except:
    pass
for i in ["CSE","MECH","IT","ECE","EEE"]:
    for j in ["I","II","III","IV"]:
        try:
            os.mkdir("data/"+i+"/"+j)
        except:
            pass
db = client["Project-X"]
collection = db["studentDet"]



@app.route("/")
def index():
    return render_template("login/index.html")




@app.route("/login", methods=["POST"])
def login():
    username = request.form["username"]
    password = request.form["password"]
    ad_collection=db['admin']
    for document in ad_collection.find():
        if document==ad_collection.find_one({"username":username,"password":password}):
            return redirect(url_for("adminhome",username=username))
    result = collection.find_one({"key": "value"})
    query={"Roll No":username}
    resultt=collection.find_one(query)
    print(resultt)
    result=""
    try:
        result=str(resultt["Reg No"])
    except:
        pass
    if password==result:
        print("done")
        return redirect(url_for("overview",username=username))
    else:
        error = "Invalid username or password. Please try again."
        return render_template("login/index.html", error=error)



@app.route("/adminhome/<username>", methods=["GET","POST"])
def adminhome(username):
    if request.method == "POST":
        year = request.form["year"]
        dept = request.form["dept"]
        dept=dept.upper()
        print(year, dept)
        print(request.files)
        if 1 :
            print("done")
            file = request.files['file']
            dirlist=os.listdir()
            dielen=len(dirlist)
            dielen=str(dielen)
            file.save("data/" + dept + "/"+ year+"/" + dielen+".xlsx")
            dir="data/" + dept + "/" +year+"/"+ dielen+".xlsx"
            init_db(dir,year)
        return render_template("adminhome/index.html", username=username)
    else:
        # Handle GET request logic (if needed)
        return render_template("adminhome/index.html", username=username)



@app.route("/adminreport/<username>",methods=["GET","POST"])
def adminreport(username):

    return render_template("adminreport/index.html",username=username)



@app.route("/adreport",methods=["GET","POST"])
def adreport():

    regno = request.form["regno"]
    report(regno)
    username=regno
    return report(regno)
    #return render_template("/report/<username>",username=username)


@app.route('/download',methods=["POST"])
def download():
    db = client["Project-X"]
    collection = db["studentDet"]
    username=request.form["regno"]
    query={"Roll No":username}
    resultt=collection.find_one(query)
    yr=resultt["Yr"]
    if "CS" in username:
        username="CSE"
    elif "IT" in username:
        username="IT"
    elif "ME" in username:
        username="MECH"
    elif "EC" in username:
        username="ECE"
    elif "EE" in username:
        username="EEE"
    list=os.listdir('data/'+username+"/"+yr)
    for i in list:
        return send_file('data/'+username+"/"+yr+"/"+i, as_attachment=True)

@app.route("/adminupdate/<username>")
def adminupdate(username):
    return render_template("adminupdate/index.html",username=username)



@app.route("/overview/<username>" )
def overview(username):
    query={"Roll No":username}
    resultt=collection.find_one(query)
    perc=resultt["attendance"]
    print(perc)
    perc_ring=perc*3.6
    perc="{:.2f}".format(perc)
    atte_per=resultt["atte_per"]
    total_per=resultt["total_per"]
    return(render_template("overview/overview.html",username=username,perc=perc,perc_ring=perc_ring,atte_per=atte_per,total_per=total_per))



@app.route("/profile/<username>")
def profile(username):
    query={"Roll No":username}
    resultt=collection.find_one(query)
    print(resultt)
    name=resultt["name"]
    print(name)
    roll=resultt["Roll No"]
    reg=resultt["Reg No"]
    yr=resultt["Yr"]
    age=resultt["Age"]
    sem=resultt["Sem"]
    dob=resultt["DOB"]
    email=resultt["Email"]
    cls=resultt["Cls"]
    ph=resultt["phone"]
    fs = gridfs.GridFS(db)
    file_cursor = fs.find_one({'filename': username})
    if file_cursor is not None:
        image_data = file_cursor.read()
        encoded_image = 'data:image/jpeg;base64,' + base64.b64encode(image_data).decode('utf-8')
    else:
        default_image_path = 'static/img/default_profile.jpg'
        encoded_image = 'data:image/jpeg;base64,' + base64.b64encode(open(default_image_path, 'rb').read()).decode('utf-8')
    return render_template("profile/index.html",username=username,result=resultt,name=name,roll=roll,reg=reg,yr=yr,age=age,sem=sem,dob=dob,email=email,cls=cls,ph=ph,image=encoded_image)
    


@app.route('/upload/<username>', methods=['POST'])
def upload(username):   
    file = request.files['image']
    fs = gridfs.GridFS(db)
    image = fs.find_one({'filename': username})
    if image:
        fs.delete(image._id)
        print("deleted")
    image_id = fs.put(file, filename=os.path.basename(username))
    print("added")
    return 'File uploaded successfully'




@app.route("/report/<username>")
def report(username):
    db = client["Project-X"]
    collection = db["studentDet"]
    query={"Roll No":username}
    resultt=collection.find_one(query)
    yr=resultt["Yr"]
    print(yr)
    try:
        os.mkdir("report_data/"+username)
    except:
        pass

    if "CS" in username:
        list=os.listdir("data/CSE/"+yr)
        dep="CSE"
    elif "IT" in username:
        list=os.listdir("data/IT/"+yr)
        dep="IT"
    elif "ME" in username:
        list=os.listdir("data/MECH/"+yr)
        dep="MECH"
    elif "EC" in username:
        list=os.listdir("data/ECE/"+yr)
        dep="ECE"
    elif "EE" in username:
        list=os.listdir("data/EEE/"+yr)
        dep="EEE"
    print(list)
    for i in list:
        print(i)
        data = pd.read_excel("data/"+dep+"/"+yr+"/"+i)
        print("data/"+dep+"/"+yr+"/"+i)
        data.to_csv('report_data/'+username+'/output.csv', index=False)
        with open('report_data/'+username+"/output.csv", mode='r') as file:
            csv_reader = csv.reader(file)
            list=[]
            for row in csv_reader:
                if "STUDENTS ATTENDANCE" in row:
                    i=row.index("STUDENTS ATTENDANCE")
                    row[i]=""
                if "DATE" in row:
                    list.append(row[4:])
                if "DAY ORDER" in row:
                    i=row.index("DAY ORDER")
                    row[i]="DAY"
                    list.append(row[4:])
                if username in row:
                    list.append(row[4:])
        fname="report_data/" + username + "/"+username+".csv"
        with open(fname,"w") as file:
            writer=csv.writer(file)
            for i in list:
                writer.writerow(i)
        data1 = pd.read_csv("report_data/"+username+"/"+username+".csv")
        for i in range(54):
            a="Unnamed: "+str(i)
            data1.rename({a:i}, axis="columns", inplace=True)
        data1 = data1.dropna(axis=1, how='all')
        data1.at[3, data1.columns[1]] = None
    output_excel_file = "report_data/"+username+'/report.xlsx'
    with pd.ExcelWriter(output_excel_file, engine='xlsxwriter') as writer:
        data1.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        for idx, col in enumerate(data1.columns):
            max_length = max(data1[col].astype(str).apply(len).max(), len(str(col)))
            worksheet.set_column(idx, idx, max_length)
    cleanup=os.listdir("report_data/"+username+"/")
    print(cleanup)
    for i in cleanup:
        if i!="report.xlsx":
            os.remove("report_data/"+username+"/"+i) 
    return send_file('report_data/'+username+'/report.xlsx', as_attachment=True)


@app.route("/welcome")
def welcome():
    return "Welcome to the protected area!"



if __name__ == "__main__":
    app.run(debug=True)
