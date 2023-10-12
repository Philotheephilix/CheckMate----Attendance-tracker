import pymongo
import openpyxl as xl
xlsx=xl.load_workbook("data/attendance.xlsx",data_only=True)
sheet=xlsx["I_CSE-B"]
column=0
last_row = None
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=4, column=row).value is not None:
        last_row = row
reg='D'
ap='BA'
tp='BB'
userlist=[]
AP = xl.utils.column_index_from_string(ap)
print(AP)
TP = xl.utils.column_index_from_string(tp)
print(sheet.cell(10,AP).value)
print(last_row)
for row in range(1, sheet.max_row + 1):
    reg_no = sheet[reg + str(row)].value
    attended=sheet.cell(row,AP).value
    total=sheet.cell(row,TP).value
    _id=row
    
    if reg_no and _id !=None and isinstance(reg_no, int):
        perc=(attended/total)*100
        user_dict = {
            "_id": _id,
            "Reg No": reg_no,
            "attendance": perc
        }
        userlist.append(user_dict)
print(userlist)
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["mydatabase"]
collection = db["studentattend"]
collection.insert_many(userlist)